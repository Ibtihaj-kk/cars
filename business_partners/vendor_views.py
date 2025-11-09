"""
Vendor-specific views for part management.
Provides vendors with comprehensive access to manage their parts with all Excel fields.
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg, F
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import json
import csv
from io import StringIO

from .models import BusinessPartner, VendorProfile, ReorderNotification
from .forms import (
    VendorPartForm, 
    VendorPartSearchForm, 
    VendorPartBulkUpdateForm,
    VendorPartBulkImportForm,
    VendorPartExportForm
)
from parts.models import Part, Category, Brand
from .permissions import (
    vendor_required, 
    vendor_part_owner_required, 
    get_vendor_profile,
    user_has_vendor_access
)


def get_vendor_profile_legacy(user):
    """Legacy function - use get_vendor_profile from permissions instead"""
    return get_vendor_profile(user)


@vendor_required
def vendor_dashboard(request):
    """
    Main vendor dashboard with overview of parts, orders, and analytics.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    # Get vendor's business partner
    business_partner = vendor_profile.business_partner
    
    # Get parts statistics
    parts_queryset = Part.objects.filter(vendor=business_partner)
    total_parts = parts_queryset.count()
    active_parts = parts_queryset.filter(is_active=True).count()
    featured_parts = parts_queryset.filter(is_featured=True).count()
    out_of_stock = parts_queryset.filter(quantity=0, is_active=True).count()
    low_stock = parts_queryset.filter(quantity__gt=0, quantity__lte=10, is_active=True).count()
    
    # Enhanced inventory statistics
    below_safety_stock = parts_queryset.filter(
        safety_stock__isnull=False,
        quantity__lt=F('safety_stock'),
        is_active=True
    ).count()
    
    # Calculate reorder recommendations
    needs_reorder = parts_queryset.filter(
        Q(quantity=0) |  # Out of stock
        Q(quantity__gt=0, quantity__lte=10) |  # Low stock
        Q(safety_stock__isnull=False, quantity__lt=F('safety_stock')),  # Below safety stock
        is_active=True
    ).count()
    
    # Get recent parts
    recent_parts = parts_queryset.order_by('-created_at')[:5]
    
    # Get top categories
    top_categories = (
        parts_queryset
        .values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    # Calculate total inventory value (corrected calculation)
    total_value = parts_queryset.aggregate(
        total=Sum(F('price') * F('quantity'))
    )['total'] or 0
    
    # Get average price
    avg_price = parts_queryset.aggregate(avg=Avg('price'))['avg'] or 0
    
    # Get inventory health metrics
    total_active_parts = parts_queryset.filter(is_active=True).count()
    healthy_stock_count = total_active_parts - out_of_stock - low_stock - below_safety_stock
    
    # Calculate inventory health percentage
    inventory_health = (healthy_stock_count / total_active_parts * 100) if total_active_parts > 0 else 100
    
    # Determine health status text
    if inventory_health >= 95:
        health_status = 'Excellent'
    elif inventory_health >= 85:
        health_status = 'Good'
    elif inventory_health >= 70:
        health_status = 'Fair'
    else:
        health_status = 'Poor'
    
    # Get recent reorder notifications
    recent_notifications = ReorderNotification.objects.filter(
        vendor=business_partner,
        status__in=['pending', 'acknowledged']
    ).select_related('part').order_by('-created_at')[:5]
    
    # Get critical notifications count
    critical_notifications = ReorderNotification.objects.filter(
        vendor=business_partner,
        priority='critical',
        status__in=['pending', 'acknowledged']
    ).count()
    
    # Get overdue notifications (older than 7 days)
    overdue_notifications = ReorderNotification.objects.filter(
        vendor=business_partner,
        status='pending',
        created_at__lt=timezone.now() - timedelta(days=7)
    ).count()
    
    context = {
        'vendor_profile': vendor_profile,
        'business_partner': business_partner,
        'stats': {
            'total_parts': total_parts,
            'active_parts': active_parts,
            'featured_parts': featured_parts,
            'out_of_stock': out_of_stock,
            'low_stock': low_stock,
            'below_safety_stock': below_safety_stock,
            'needs_reorder': needs_reorder,
            'total_value': total_value,
            'avg_price': avg_price,
            'inventory_health': inventory_health,
            'inventory_health_percentage': round(inventory_health, 1),
            'health_status': health_status,
            'healthy_stock_count': healthy_stock_count,
            'critical_notifications': critical_notifications,
            'overdue_notifications': overdue_notifications,
        },
        'recent_notifications': recent_notifications,
        'recent_parts': recent_parts,
        'top_categories': top_categories,
    }
    
    return render(request, 'business_partners/vendor_dashboard.html', context)


@vendor_required
def vendor_parts_list(request):
    """
    List all parts for the vendor with advanced filtering and search.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get all vendor's parts
    parts_queryset = Part.objects.filter(vendor=business_partner)
    
    # Initialize search form
    search_form = VendorPartSearchForm(request.GET or None)
    
    # Apply filters
    if search_form.is_valid():
        # Basic search
        search_query = search_form.cleaned_data.get('search')
        if search_query:
            parts_queryset = parts_queryset.filter(
                Q(parts_number__icontains=search_query) |
                Q(material_description__icontains=search_query) |
                Q(material_description_ar__icontains=search_query) |
                Q(manufacturer_part_number__icontains=search_query) |
                Q(manufacturer_oem_number__icontains=search_query)
            )
        
        # Category filter
        category = search_form.cleaned_data.get('category')
        if category:
            parts_queryset = parts_queryset.filter(category=category)
        
        # Brand filter
        brand = search_form.cleaned_data.get('brand')
        if brand:
            parts_queryset = parts_queryset.filter(brand=brand)
        
        # Status filters
        is_active = search_form.cleaned_data.get('is_active')
        if is_active == 'true':
            parts_queryset = parts_queryset.filter(is_active=True)
        elif is_active == 'false':
            parts_queryset = parts_queryset.filter(is_active=False)
        
        is_featured = search_form.cleaned_data.get('is_featured')
        if is_featured == 'true':
            parts_queryset = parts_queryset.filter(is_featured=True)
        elif is_featured == 'false':
            parts_queryset = parts_queryset.filter(is_featured=False)
        
        # Stock status filter
        stock_status = search_form.cleaned_data.get('stock_status')
        if stock_status == 'in_stock':
            parts_queryset = parts_queryset.filter(quantity__gt=10)
        elif stock_status == 'low_stock':
            parts_queryset = parts_queryset.filter(quantity__gt=0, quantity__lte=10)
        elif stock_status == 'out_of_stock':
            parts_queryset = parts_queryset.filter(quantity=0)
        
        # Price range
        min_price = search_form.cleaned_data.get('min_price')
        if min_price:
            parts_queryset = parts_queryset.filter(price__gte=min_price)
        
        max_price = search_form.cleaned_data.get('max_price')
        if max_price:
            parts_queryset = parts_queryset.filter(price__lte=max_price)
        
        # Vendor-specific filters
        material_type = search_form.cleaned_data.get('material_type')
        if material_type:
            parts_queryset = parts_queryset.filter(material_type__icontains=material_type)
        
        plant = search_form.cleaned_data.get('plant')
        if plant:
            parts_queryset = parts_queryset.filter(plant__icontains=plant)
        
        material_group = search_form.cleaned_data.get('material_group')
        if material_group:
            parts_queryset = parts_queryset.filter(material_group__icontains=material_group)
        
        abc_indicator = search_form.cleaned_data.get('abc_indicator')
        if abc_indicator:
            parts_queryset = parts_queryset.filter(abc_indicator=abc_indicator)
        
        # Sorting
        sort_by = search_form.cleaned_data.get('sort_by')
        if sort_by:
            parts_queryset = parts_queryset.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(parts_queryset, 25)  # Show 25 parts per page
    page_number = request.GET.get('page')
    parts = paginator.get_page(page_number)
    
    context = {
        'parts': parts,
        'search_form': search_form,
        'vendor_profile': vendor_profile,
        'total_parts': parts_queryset.count(),
    }
    
    return render(request, 'business_partners/vendor_parts_list.html', context)


@vendor_required
def vendor_part_create(request):
    """
    Create a new part for the vendor with access to all Excel fields.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    if request.method == 'POST':
        form = VendorPartForm(request.POST, request.FILES, vendor=business_partner)
        if form.is_valid():
            part = form.save()
            messages.success(request, f'Part "{part.parts_number}" created successfully.')
            return redirect('vendor_parts_list')
    else:
        form = VendorPartForm(vendor=business_partner)
    
    context = {
        'form': form,
        'vendor_profile': vendor_profile,
        'action': 'Create',
    }
    
    return render(request, 'business_partners/vendor_part_form.html', context)


@vendor_part_owner_required
def vendor_part_edit(request, part_id):
    """
    Edit an existing part for the vendor with access to all Excel fields.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get the part and ensure it belongs to this vendor
    part = get_object_or_404(Part, id=part_id, vendor=business_partner)
    
    if request.method == 'POST':
        form = VendorPartForm(request.POST, request.FILES, instance=part, vendor=business_partner)
        if form.is_valid():
            part = form.save()
            messages.success(request, f'Part "{part.parts_number}" updated successfully.')
            return redirect('vendor_parts_list')
    else:
        form = VendorPartForm(instance=part, vendor=business_partner)
    
    context = {
        'form': form,
        'part': part,
        'vendor_profile': vendor_profile,
        'action': 'Edit',
    }
    
    return render(request, 'business_partners/vendor_part_form.html', context)


@vendor_part_owner_required
def vendor_part_detail(request, part_id):
    """
    View detailed information about a specific part.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get the part and ensure it belongs to this vendor
    part = get_object_or_404(Part, id=part_id, vendor=business_partner)
    
    # Get categorized fields based on user permissions
    visible_fields = part.get_fields_for_user(request.user)
    can_edit = part.can_user_edit(request.user)
    
    context = {
        'part': part,
        'visible_fields': visible_fields,
        'can_edit': can_edit,
        'vendor_profile': vendor_profile,
    }
    
    return render(request, 'business_partners/vendor_part_detail.html', context)


@vendor_part_owner_required
@require_http_methods(["POST"])
def vendor_part_delete(request, part_id):
    """
    Delete a part (AJAX endpoint).
    Permission checking is handled by the decorator.
    """
    try:
        part = get_object_or_404(Part, id=part_id)
        part_number = part.parts_number
        part.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Part "{part_number}" deleted successfully.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@vendor_required
def vendor_parts_bulk_update(request):
    """
    Bulk update multiple parts at once.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    if request.method == 'POST':
        form = VendorPartBulkUpdateForm(request.POST)
        part_ids = request.POST.getlist('selected_parts')
        
        if form.is_valid() and part_ids:
            # Get selected parts that belong to this vendor
            parts = Part.objects.filter(id__in=part_ids, vendor=business_partner)
            
            if not parts.exists():
                messages.error(request, 'No valid parts selected.')
                return redirect('vendor_parts_list')
            
            updated_count = 0
            
            with transaction.atomic():
                for part in parts:
                    updated = False
                    
                    # Price adjustments
                    price_type = form.cleaned_data.get('price_adjustment_type')
                    price_value = form.cleaned_data.get('price_adjustment_value')
                    
                    if price_type and price_value:
                        if price_type == 'percentage_increase':
                            part.price = part.price * (1 + price_value / 100)
                            updated = True
                        elif price_type == 'percentage_decrease':
                            part.price = part.price * (1 - price_value / 100)
                            updated = True
                        elif price_type == 'fixed_amount_increase':
                            part.price = part.price + price_value
                            updated = True
                        elif price_type == 'fixed_amount_decrease':
                            part.price = max(Decimal('0.01'), part.price - price_value)
                            updated = True
                        elif price_type == 'set_price':
                            part.price = price_value
                            updated = True
                    
                    # Quantity adjustments
                    quantity_type = form.cleaned_data.get('quantity_adjustment_type')
                    quantity_value = form.cleaned_data.get('quantity_adjustment_value')
                    
                    if quantity_type and quantity_value is not None:
                        if quantity_type == 'add':
                            part.quantity = part.quantity + quantity_value
                            updated = True
                        elif quantity_type == 'subtract':
                            part.quantity = max(0, part.quantity - quantity_value)
                            updated = True
                        elif quantity_type == 'set':
                            part.quantity = quantity_value
                            updated = True
                    
                    # Status updates
                    active_status = form.cleaned_data.get('set_active_status')
                    if active_status:
                        part.is_active = active_status == 'true'
                        updated = True
                    
                    featured_status = form.cleaned_data.get('set_featured_status')
                    if featured_status:
                        part.is_featured = featured_status == 'true'
                        updated = True
                    
                    # Vendor-specific updates
                    material_type = form.cleaned_data.get('update_material_type')
                    if material_type:
                        part.material_type = material_type
                        updated = True
                    
                    plant = form.cleaned_data.get('update_plant')
                    if plant:
                        part.plant = plant
                        updated = True
                    
                    material_group = form.cleaned_data.get('update_material_group')
                    if material_group:
                        part.material_group = material_group
                        updated = True
                    
                    if updated:
                        part.save()
                        updated_count += 1
            
            messages.success(request, f'Successfully updated {updated_count} parts.')
            return redirect('vendor_parts_list')
        else:
            if not part_ids:
                messages.error(request, 'No parts selected for bulk update.')
            else:
                messages.error(request, 'Please correct the errors in the form.')
    else:
        form = VendorPartBulkUpdateForm()
    
    context = {
        'form': form,
        'vendor_profile': vendor_profile,
    }
    
    return render(request, 'business_partners/vendor_parts_bulk_update.html', context)


@login_required
def vendor_parts_export_csv(request):
    """
    Export vendor's parts to CSV with all Excel fields.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get all vendor's parts
    parts = Part.objects.filter(vendor=business_partner)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="vendor_parts_{business_partner.name}.csv"'
    
    writer = csv.writer(response)
    
    # Write header row with all Excel fields
    header = [
        'Parts Number', 'Material Description', 'Material Description (AR)',
        'Base Unit of Measure', 'Gross Weight', 'Net Weight', 'Size Dimensions',
        'Manufacturer Part Number', 'Manufacturer OEM Number',
        'Material Type', 'Plant', 'Storage Location', 'Warehouse Number',
        'Material Group', 'Division', 'Minimum Order Quantity', 'Old Material Number',
        'Expiration XCHPF', 'External Material Group', 'ABC Indicator',
        'Safety Stock', 'Minimum Safety Stock', 'Planned Delivery Time (Days)',
        'Goods Receipt Processing Time (Days)', 'Valuation Class', 'Price Unit (PEINH)',
        'Moving Average Price', 'Category', 'Brand', 'Price', 'Quantity',
        'Warranty Period', 'Is Active', 'Is Featured', 'Created At', 'Updated At'
    ]
    writer.writerow(header)
    
    # Write data rows
    for part in parts:
        row = [
            part.parts_number or '',
            part.material_description or '',
            part.material_description_ar or '',
            part.base_unit_of_measure or '',
            part.gross_weight or '',
            part.net_weight or '',
            part.size_dimensions or '',
            part.manufacturer_part_number or '',
            part.manufacturer_oem_number or '',
            part.material_type or '',
            part.plant or '',
            part.storage_location or '',
            part.warehouse_number or '',
            part.material_group or '',
            part.division or '',
            part.minimum_order_quantity or '',
            part.old_material_number or '',
            part.expiration_xchpf or '',
            part.external_material_group or '',
            part.abc_indicator or '',
            part.safety_stock or '',
            part.minimum_safety_stock or '',
            part.planned_delivery_time_days or '',
            part.goods_receipt_processing_time_days or '',
            part.valuation_class or '',
            part.price_unit_peinh or '',
            part.moving_average_price or '',
            part.category.name if part.category else '',
            part.brand.name if part.brand else '',
            part.price or '',
            part.quantity or '',
            part.warranty_period or '',
            'Yes' if part.is_active else 'No',
            'Yes' if part.is_featured else 'No',
            part.created_at.strftime('%Y-%m-%d %H:%M:%S') if part.created_at else '',
            part.updated_at.strftime('%Y-%m-%d %H:%M:%S') if part.updated_at else '',
        ]
        writer.writerow(row)
    
    return response


@vendor_required
@require_http_methods(["POST"])
def vendor_parts_bulk_action(request):
    """
    Handle AJAX bulk actions on vendor parts from the parts list.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        return JsonResponse({'success': False, 'error': 'Access denied'})
    
    business_partner = vendor_profile.business_partner
    
    try:
        part_ids = request.POST.getlist('part_ids')
        action = request.POST.get('action')
        
        if not part_ids:
            return JsonResponse({'success': False, 'error': 'No parts selected'})
        
        # Get selected parts that belong to this vendor
        parts = Part.objects.filter(id__in=part_ids, vendor=business_partner)
        
        if not parts.exists():
            return JsonResponse({'success': False, 'error': 'No valid parts found'})
        
        updated_count = 0
        
        with transaction.atomic():
            if action == 'activate':
                parts.update(is_active=True)
                updated_count = parts.count()
                message = f'Activated {updated_count} parts'
                
            elif action == 'deactivate':
                parts.update(is_active=False)
                updated_count = parts.count()
                message = f'Deactivated {updated_count} parts'
                
            elif action == 'feature':
                parts.update(is_featured=True)
                updated_count = parts.count()
                message = f'Featured {updated_count} parts'
                
            elif action == 'unfeature':
                parts.update(is_featured=False)
                updated_count = parts.count()
                message = f'Removed featured status from {updated_count} parts'
                
            elif action == 'delete':
                updated_count = parts.count()
                parts.delete()
                message = f'Deleted {updated_count} parts'
                
            elif action == 'update_price':
                # Price update with percentage
                percentage = float(request.POST.get('percentage', 0))
                if percentage != 0:
                    for part in parts:
                        adjustment_factor = Decimal(str(1 + (percentage / 100)))
                        part.price = part.price * adjustment_factor
                        part.save()
                        updated_count += 1
                    message = f'Updated prices for {updated_count} parts by {percentage}%'
                else:
                    return JsonResponse({'success': False, 'error': 'Invalid percentage value'})
                    
            elif action == 'update_stock':
                # Stock update
                stock_action = request.POST.get('stock_action', 'set')
                quantity = int(request.POST.get('quantity', 0))
                
                for part in parts:
                    if stock_action == 'set':
                        part.quantity = quantity
                    elif stock_action == 'add':
                        part.quantity += quantity
                    elif stock_action == 'subtract':
                        part.quantity = max(0, part.quantity - quantity)
                    part.save()
                    updated_count += 1
                    
                message = f'Updated stock for {updated_count} parts'
                
            else:
                return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        return JsonResponse({
            'success': True,
            'message': message,
            'updated_count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def vendor_store_front(request, vendor_slug):
    """
    Public view for vendor store front - displays vendor information and their parts.
    Uses the same design patterns as the main parts catalog.
    """
    from parts.models import Part
    from parts.views import get_cached_categories, get_cached_brands
    from parts.forms import PartSearchForm
    from django.core.paginator import Paginator
    
    # Get vendor by slug
    try:
        vendor = BusinessPartner.objects.select_related('vendor_profile').get(
            slug=vendor_slug,
            roles__role_type='vendor',
            status='active'
        )
    except BusinessPartner.DoesNotExist:
        messages.error(request, 'Vendor not found.')
        return redirect('parts:part_list')
    
    # Get vendor's parts with same filtering as main catalog
    queryset = Part.objects.filter(
        vendor=vendor,
        is_active=True
    ).select_related(
        'category', 'brand', 'dealer', 'inventory'
    ).prefetch_related(
        'reviews', 'cart_items'
    ).annotate(
        annotated_review_count=Count('reviews', filter=Q(reviews__is_approved=True)),
        avg_rating=Avg('reviews__rating', filter=Q(reviews__is_approved=True)),
        total_sales=Count('order_items')
    )
    
    # Apply same filters as main parts list
    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__name__icontains=search_query) |
            Q(parts_number__icontains=search_query) |
            Q(material_description__icontains=search_query) |
            Q(manufacturer_part_number__icontains=search_query)
        )
    
    # Category filter
    category_id = request.GET.get('category')
    if category_id:
        queryset = queryset.filter(category_id=category_id)
    
    categories = request.GET.getlist('categories')
    if categories:
        queryset = queryset.filter(category_id__in=categories)
    
    # Brand filter
    brand_id = request.GET.get('brand')
    if brand_id:
        queryset = queryset.filter(brand_id=brand_id)
    
    brands = request.GET.getlist('brands')
    if brands:
        queryset = queryset.filter(brand_id__in=brands)
    
    # Price range filter
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    if min_price:
        queryset = queryset.filter(price__gte=min_price)
    if max_price:
        queryset = queryset.filter(price__lte=max_price)
    
    # Stock filter
    stock_filter = request.GET.get('stock_availability')
    if stock_filter == 'in_stock':
        queryset = queryset.filter(quantity__gt=0)
    elif stock_filter == 'pre_order':
        queryset = queryset.filter(quantity=0, is_active=True)
    
    # Legacy stock filters
    in_stock = request.GET.get('in_stock')
    if in_stock == 'true':
        queryset = queryset.filter(quantity__gt=0)
    
    # Featured filter
    featured = request.GET.get('featured')
    if featured == 'true':
        queryset = queryset.filter(is_featured=True)
    
    # Minimum rating filter
    min_rating = request.GET.get('min_rating')
    if min_rating:
        queryset = queryset.filter(avg_rating__gte=min_rating)
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    if sort_by in ['name', '-name', 'price', '-price', 'created_at', '-created_at']:
        queryset = queryset.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(queryset, 12)
    page_number = request.GET.get('page')
    parts = paginator.get_page(page_number)
    
    # HTMX partial response for filtering
    if request.headers.get('HX-Request'):
        return render(request, 'parts/partials/parts_results.html', {'parts': parts})
    
    context = {
        'vendor': vendor,
        'parts': parts,
        'search_form': PartSearchForm(request.GET),
        'categories': get_cached_categories(),
        'brands': get_cached_brands(),
        'total_parts': queryset.count(),
        'view_type': 'vendor_store',  # For template conditionals
    }
    
    return render(request, 'business_partners/vendor_store_front.html', context)


@login_required
def vendor_reorder_notifications(request):
    """
    Display reorder notifications for the vendor.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get filter parameters
    status_filter = request.GET.get('status', 'pending')
    priority_filter = request.GET.get('priority', '')
    
    # Base queryset
    notifications = ReorderNotification.objects.filter(
        vendor=business_partner
    ).select_related('part', 'acknowledged_by').order_by('-urgency_score', '-created_at')
    
    # Apply filters
    if status_filter and status_filter != 'all':
        notifications = notifications.filter(status=status_filter)
    
    if priority_filter and priority_filter != 'all':
        notifications = notifications.filter(priority=priority_filter)
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get statistics
    stats = {
        'total': ReorderNotification.objects.filter(vendor=business_partner).count(),
        'pending': ReorderNotification.objects.filter(vendor=business_partner, status='pending').count(),
        'acknowledged': ReorderNotification.objects.filter(vendor=business_partner, status='acknowledged').count(),
        'ordered': ReorderNotification.objects.filter(vendor=business_partner, status='ordered').count(),
        'critical': ReorderNotification.objects.filter(vendor=business_partner, priority='critical', status__in=['pending', 'acknowledged']).count(),
        'overdue': ReorderNotification.objects.filter(
            vendor=business_partner,
            status='pending',
            created_at__lt=timezone.now() - timedelta(days=7)
        ).count(),
    }
    
    context = {
        'vendor_profile': vendor_profile,
        'notifications': page_obj,
        'stats': stats,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'status_choices': ReorderNotification.STATUS_CHOICES,
        'priority_choices': ReorderNotification.PRIORITY_CHOICES,
    }
    
    return render(request, 'business_partners/vendor_reorder_notifications.html', context)


@login_required
@require_http_methods(["POST"])
def vendor_reorder_notification_action(request, notification_id):
    """
    Handle actions on reorder notifications (acknowledge, mark ordered, dismiss, etc.).
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        return JsonResponse({'success': False, 'error': 'Access denied'})
    
    business_partner = vendor_profile.business_partner
    
    try:
        notification = ReorderNotification.objects.get(
            id=notification_id,
            vendor=business_partner
        )
    except ReorderNotification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'})
    
    action = request.POST.get('action')
    
    if action == 'acknowledge':
        notification.acknowledge(request.user)
        return JsonResponse({
            'success': True,
            'message': 'Notification acknowledged',
            'new_status': notification.get_status_display()
        })
    
    elif action == 'mark_ordered':
        order_reference = request.POST.get('order_reference', '')
        expected_delivery = request.POST.get('expected_delivery')
        
        # Parse expected delivery date if provided
        expected_delivery_date = None
        if expected_delivery:
            try:
                from datetime import datetime
                expected_delivery_date = datetime.strptime(expected_delivery, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid date format'})
        
        notification.mark_ordered(order_reference, expected_delivery_date)
        return JsonResponse({
            'success': True,
            'message': 'Notification marked as ordered',
            'new_status': notification.get_status_display()
        })
    
    elif action == 'complete':
        notification.complete()
        return JsonResponse({
            'success': True,
            'message': 'Notification completed',
            'new_status': notification.get_status_display()
        })
    
    elif action == 'dismiss':
        notification.dismiss()
        return JsonResponse({
            'success': True,
            'message': 'Notification dismissed',
            'new_status': notification.get_status_display()
        })
    
    else:
        return JsonResponse({'success': False, 'error': 'Invalid action'})


@login_required
@require_http_methods(["POST"])
def vendor_bulk_reorder_action(request):
    """
    Handle bulk actions on multiple reorder notifications.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        return JsonResponse({'success': False, 'error': 'Access denied'})
    
    business_partner = vendor_profile.business_partner
    
    notification_ids = request.POST.getlist('notification_ids')
    action = request.POST.get('action')
    
    if not notification_ids:
        return JsonResponse({'success': False, 'error': 'No notifications selected'})
    
    # Get notifications
    notifications = ReorderNotification.objects.filter(
        id__in=notification_ids,
        vendor=business_partner
    )
    
    if not notifications.exists():
        return JsonResponse({'success': False, 'error': 'No valid notifications found'})
    
    count = 0
    
    if action == 'acknowledge':
        for notification in notifications:
            if notification.status == 'pending':
                notification.acknowledge(request.user)
                count += 1
    
    elif action == 'dismiss':
        for notification in notifications:
            if notification.status in ['pending', 'acknowledged']:
                notification.dismiss()
                count += 1
    
    else:
        return JsonResponse({'success': False, 'error': 'Invalid action'})
    
    return JsonResponse({
        'success': True,
        'message': f'{count} notifications {action}d successfully'
    })


@login_required
def vendor_reorder_notification_detail(request, notification_id):
    """
    Display detailed view of a reorder notification.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    try:
        notification = ReorderNotification.objects.select_related(
            'part', 'part__category', 'part__brand', 'acknowledged_by'
        ).get(
            id=notification_id,
            vendor=business_partner
        )
    except ReorderNotification.DoesNotExist:
        messages.error(request, 'Notification not found.')
        return redirect('vendor_reorder_notifications')
    
    # Get part inventory information
    inventory = getattr(notification.part, 'inventory', None)
    
    # Get recent notifications for this part
    recent_notifications = ReorderNotification.objects.filter(
        part=notification.part,
        vendor=business_partner
    ).exclude(id=notification.id).order_by('-created_at')[:5]
    
    context = {
        'vendor_profile': vendor_profile,
        'notification': notification,
        'inventory': inventory,
        'recent_notifications': recent_notifications,
    }
    
    return render(request, 'business_partners/vendor_reorder_notification_detail.html', context)


@login_required
def vendor_inventory_alerts(request):
    """
    Show inventory alerts for low stock and out of stock items.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    # Get parts with stock issues
    out_of_stock = Part.objects.filter(vendor=business_partner, quantity=0, is_active=True)
    low_stock = Part.objects.filter(
        vendor=business_partner, 
        quantity__gt=0, 
        quantity__lte=10, 
        is_active=True
    )
    
    # Get parts below safety stock
    below_safety_stock = Part.objects.filter(
        vendor=business_partner,
        safety_stock__isnull=False,
        quantity__lt=models.F('safety_stock'),
        is_active=True
    )
    
    context = {
        'vendor_profile': vendor_profile,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'below_safety_stock': below_safety_stock,
    }
    
    return render(request, 'business_partners/vendor_inventory_alerts.html', context)


@login_required
def vendor_parts_import(request):
    """
    Handle bulk import of vendor parts from CSV/Excel files.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    if request.method == 'POST':
        form = VendorPartBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            import_file = form.cleaned_data['import_file']
            update_existing = form.cleaned_data['update_existing']
            validate_only = form.cleaned_data['validate_only']
            
            try:
                # Process the import file
                results = process_import_file(
                    import_file, 
                    business_partner, 
                    update_existing, 
                    validate_only
                )
                
                if validate_only:
                    messages.info(request, f'Validation complete. {results["valid_count"]} valid rows, {results["error_count"]} errors.')
                else:
                    messages.success(request, f'Import complete. {results["created_count"]} parts created, {results["updated_count"]} parts updated.')
                
                # Store results in session for display
                request.session['import_results'] = results
                return redirect('vendor_parts_import_results')
                
            except Exception as e:
                messages.error(request, f'Import failed: {str(e)}')
    else:
        form = VendorPartBulkImportForm()
    
    context = {
        'vendor_profile': vendor_profile,
        'form': form,
    }
    
    return render(request, 'business_partners/vendor_parts_import.html', context)


@login_required
def vendor_parts_import_results(request):
    """
    Display results of the bulk import operation.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    results = request.session.get('import_results', {})
    if not results:
        messages.warning(request, 'No import results found.')
        return redirect('vendor_parts_import')
    
    # Clear results from session
    if 'import_results' in request.session:
        del request.session['import_results']
    
    context = {
        'vendor_profile': vendor_profile,
        'results': results,
    }
    
    return render(request, 'business_partners/vendor_parts_import_results.html', context)


@login_required
def vendor_parts_export(request):
    """
    Handle export of vendor parts to CSV/Excel files.
    """
    vendor_profile = get_vendor_profile(request.user)
    if not vendor_profile:
        messages.error(request, 'You do not have vendor access.')
        return redirect('home')
    
    business_partner = vendor_profile.business_partner
    
    if request.method == 'POST':
        form = VendorPartExportForm(request.POST)
        if form.is_valid():
            export_format = form.cleaned_data['export_format']
            include_inactive = form.cleaned_data['include_inactive']
            include_images = form.cleaned_data['include_images']
            date_from = form.cleaned_data['date_from']
            date_to = form.cleaned_data['date_to']
            
            # Build queryset
            queryset = Part.objects.filter(vendor=business_partner)
            
            if not include_inactive:
                queryset = queryset.filter(is_active=True)
            
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            
            # Generate export file
            if export_format == 'csv':
                return generate_csv_export(queryset, include_images)
            else:
                return generate_excel_export(queryset, include_images)
    else:
        form = VendorPartExportForm()
    
    context = {
        'vendor_profile': vendor_profile,
        'form': form,
    }
    
    return render(request, 'business_partners/vendor_parts_export.html', context)


def process_import_file(import_file, business_partner, update_existing, validate_only):
    """
    Process CSV or Excel import file and create/update parts with comprehensive validation.
    """
    import csv
    import io
    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date
    from django.core.validators import URLValidator
    from django.core.exceptions import ValidationError
    from vehicles.models import VehicleVariant
    
    results = {
        'total_rows': 0,
        'valid_count': 0,
        'error_count': 0,
        'created_count': 0,
        'updated_count': 0,
        'errors': [],
        'warnings': [],
        'field_errors': {},  # Track errors by field type
        'validation_summary': {}
    }
    
    # Define field validation rules
    REQUIRED_FIELDS = ['parts_number', 'material_description', 'base_unit_of_measure', 'category_name', 'brand_name', 'price']
    
    FIELD_MAPPINGS = {
        'parts_number': 'part_number',
        'material_description': 'name',
        'material_description_ar': 'description_ar',
        'category_name': 'category',
        'brand_name': 'brand',
        'vehicle_variants': 'compatible_vehicles'
    }
    
    NUMERIC_FIELDS = {
        'price': {'min': 0, 'max': 999999.99, 'decimal_places': 2},
        'quantity': {'min': 0, 'max': 999999, 'decimal_places': 0},
        'safety_stock': {'min': 0, 'max': 999999, 'decimal_places': 0},
        'minimum_safety_stock': {'min': 0, 'max': 999999, 'decimal_places': 0},
        'reorder_point': {'min': 0, 'max': 999999, 'decimal_places': 0},
        'minimum_order_quantity': {'min': 1, 'max': 999999, 'decimal_places': 0},
        'inventory_threshold': {'min': 0, 'max': 999999, 'decimal_places': 0},
        'gross_weight': {'min': 0, 'max': 99999.999, 'decimal_places': 3},
        'net_weight': {'min': 0, 'max': 99999.999, 'decimal_places': 3},
        'planned_delivery_time_days': {'min': 0, 'max': 365, 'decimal_places': 0},
        'goods_receipt_processing_time_days': {'min': 0, 'max': 30, 'decimal_places': 0},
        'warranty_period': {'min': 0, 'max': 120, 'decimal_places': 0},
        'standard_price': {'min': 0, 'max': 999999.99, 'decimal_places': 2},
        'moving_average_price': {'min': 0, 'max': 999999.99, 'decimal_places': 2},
        'cost_price': {'min': 0, 'max': 999999.99, 'decimal_places': 2},
        'price_unit_peinh': {'min': 1, 'max': 99999, 'decimal_places': 0}
    }
    
    STRING_FIELDS = {
        'parts_number': {'max_length': 50},
        'material_description': {'max_length': 200},
        'material_description_ar': {'max_length': 200},
        'base_unit_of_measure': {'max_length': 10},
        'size_dimensions': {'max_length': 100},
        'manufacturer_part_number': {'max_length': 40},
        'manufacturer_oem_number': {'max_length': 50},
        'material_type': {'max_length': 50},
        'plant': {'max_length': 50},
        'storage_location': {'max_length': 50},
        'warehouse_number': {'max_length': 50},
        'storage_bin': {'max_length': 50},
        'material_group': {'max_length': 50},
        'division': {'max_length': 50},
        'old_material_number': {'max_length': 50},
        'external_material_group': {'max_length': 50},
        'abc_indicator': {'max_length': 1, 'choices': ['A', 'B', 'C']},
        'mrp_type': {'max_length': 10},
        'mrp_group': {'max_length': 10},
        'mrp_controller': {'max_length': 10},
        'valuation_class': {'max_length': 50},
        'price_control_indicator': {'max_length': 10},
        'storage_location_code': {'max_length': 50}
    }
    
    BOOLEAN_FIELDS = ['is_active', 'is_featured']
    DATE_FIELDS = ['expiration_xchpf']
    URL_FIELDS = ['image_url']

    def validate_field(field_name, value, row_num):
        """Validate individual field with comprehensive rules"""
        errors = []
        
        if not value or (isinstance(value, str) and not value.strip()):
            if field_name in REQUIRED_FIELDS:
                errors.append(f"Row {row_num}: {field_name} is required")
            return errors, None
        
        # Clean string value
        if isinstance(value, str):
            value = value.strip()
        
        # Validate string fields
        if field_name in STRING_FIELDS:
            rules = STRING_FIELDS[field_name]
            if len(str(value)) > rules['max_length']:
                errors.append(f"Row {row_num}: {field_name} exceeds maximum length of {rules['max_length']} characters")
            
            if 'choices' in rules and value not in rules['choices']:
                errors.append(f"Row {row_num}: {field_name} must be one of: {', '.join(rules['choices'])}")
        
        # Validate numeric fields
        elif field_name in NUMERIC_FIELDS:
            rules = NUMERIC_FIELDS[field_name]
            try:
                decimal_value = Decimal(str(value))
                if decimal_value < rules['min'] or decimal_value > rules['max']:
                    errors.append(f"Row {row_num}: {field_name} must be between {rules['min']} and {rules['max']}")
                
                # Check decimal places
                if rules['decimal_places'] == 0 and decimal_value != int(decimal_value):
                    errors.append(f"Row {row_num}: {field_name} must be a whole number")
                
                value = decimal_value
            except (InvalidOperation, ValueError):
                errors.append(f"Row {row_num}: {field_name} must be a valid number")
                return errors, None
        
        # Validate boolean fields
        elif field_name in BOOLEAN_FIELDS:
            if isinstance(value, str):
                value_lower = value.lower()
                if value_lower in ['yes', 'true', '1', 'y', 'on']:
                    value = True
                elif value_lower in ['no', 'false', '0', 'n', 'off', '']:
                    value = False
                else:
                    errors.append(f"Row {row_num}: {field_name} must be TRUE/FALSE, YES/NO, or 1/0")
                    return errors, None
        
        # Validate date fields
        elif field_name in DATE_FIELDS:
            try:
                parsed_date = parse_date(str(value))
                if not parsed_date:
                    errors.append(f"Row {row_num}: {field_name} must be in YYYY-MM-DD format")
                    return errors, None
                value = parsed_date
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: {field_name} must be a valid date in YYYY-MM-DD format")
                return errors, None
        
        # Validate URL fields
        elif field_name in URL_FIELDS:
            validator = URLValidator()
            try:
                validator(value)
            except ValidationError:
                errors.append(f"Row {row_num}: {field_name} must be a valid URL")
        
        return errors, value

    def validate_vehicle_variants(variants_str, row_num):
        """Validate vehicle variants field"""
        errors = []
        valid_variants = []
        
        if not variants_str or not variants_str.strip():
            return errors, valid_variants
        
        # Split by comma and clean
        variant_names = [v.strip() for v in variants_str.split(',') if v.strip()]
        
        for variant_name in variant_names:
            try:
                variant = VehicleVariant.objects.get(name=variant_name)
                valid_variants.append(variant)
            except VehicleVariant.DoesNotExist:
                errors.append(f"Row {row_num}: Vehicle variant '{variant_name}' not found in system")
        
        return errors, valid_variants

    try:
        # Determine file type and read data
        file_extension = import_file.name.lower().split('.')[-1]
        
        if file_extension == 'csv':
            # Handle CSV file
            file_content = import_file.read().decode('utf-8-sig')  # Handle BOM
            csv_reader = csv.DictReader(io.StringIO(file_content))
            rows = list(csv_reader)
        elif file_extension in ['xlsx', 'xls']:
            # Handle Excel file
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(import_file)
                worksheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in worksheet[1]]
                
                # Get data rows
                rows = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
            except ImportError:
                raise Exception("openpyxl library is required for Excel file processing")
        else:
            raise Exception("Unsupported file format. Please use CSV or Excel files.")
        
        results['total_rows'] = len(rows)
        
        # Initialize field error tracking
        for field in REQUIRED_FIELDS + list(NUMERIC_FIELDS.keys()) + list(STRING_FIELDS.keys()) + BOOLEAN_FIELDS + DATE_FIELDS + URL_FIELDS:
            results['field_errors'][field] = 0
        
        for row_num, row in enumerate(rows, start=2):
            row_errors = []
            row_warnings = []
            validated_data = {}
            
            try:
                # Validate all fields
                for field_name, raw_value in row.items():
                    if field_name and raw_value is not None:
                        field_errors, validated_value = validate_field(field_name, raw_value, row_num)
                        row_errors.extend(field_errors)
                        
                        if field_errors:
                            results['field_errors'][field_name] = results['field_errors'].get(field_name, 0) + len(field_errors)
                        
                        if validated_value is not None:
                            # Map field names
                            mapped_field = FIELD_MAPPINGS.get(field_name, field_name)
                            validated_data[mapped_field] = validated_value
                
                # Special validation for required fields
                for required_field in REQUIRED_FIELDS:
                    if required_field not in row or not row[required_field]:
                        row_errors.append(f"Row {row_num}: {required_field} is required")
                        results['field_errors'][required_field] = results['field_errors'].get(required_field, 0) + 1
                
                # Validate vehicle variants
                if 'vehicle_variants' in row:
                    variant_errors, valid_variants = validate_vehicle_variants(row['vehicle_variants'], row_num)
                    row_errors.extend(variant_errors)
                    if valid_variants:
                        validated_data['compatible_vehicles'] = valid_variants
                
                # Validate category and brand existence
                if 'category_name' in validated_data:
                    try:
                        category = Category.objects.get(name=validated_data['category_name'])
                        validated_data['category'] = category
                        del validated_data['category_name']
                    except Category.DoesNotExist:
                        row_errors.append(f"Row {row_num}: Category '{validated_data['category_name']}' not found")
                        results['field_errors']['category_name'] = results['field_errors'].get('category_name', 0) + 1
                
                if 'brand_name' in validated_data:
                    try:
                        brand = Brand.objects.get(name=validated_data['brand_name'])
                        validated_data['brand'] = brand
                        del validated_data['brand_name']
                    except Brand.DoesNotExist:
                        row_errors.append(f"Row {row_num}: Brand '{validated_data['brand_name']}' not found")
                        results['field_errors']['brand_name'] = results['field_errors'].get('brand_name', 0) + 1
                
                # Check for duplicate part number
                if 'part_number' in validated_data:
                    existing_part = None
                    try:
                        existing_part = Part.objects.get(
                            part_number=validated_data['part_number'],
                            vendor=business_partner
                        )
                    except Part.DoesNotExist:
                        pass
                    
                    if existing_part and not update_existing:
                        row_warnings.append(f"Row {row_num}: Part {validated_data['part_number']} already exists (skipped)")
                        continue
                
                # Business logic validations
                if 'safety_stock' in validated_data and 'quantity' in validated_data:
                    if validated_data['safety_stock'] > validated_data['quantity']:
                        row_warnings.append(f"Row {row_num}: Safety stock is higher than current quantity")
                
                if 'reorder_point' in validated_data and 'safety_stock' in validated_data:
                    if validated_data['reorder_point'] < validated_data['safety_stock']:
                        row_warnings.append(f"Row {row_num}: Reorder point should be higher than safety stock")
                
                if 'cost_price' in validated_data and 'price' in validated_data:
                    if validated_data['cost_price'] > validated_data['price']:
                        row_warnings.append(f"Row {row_num}: Cost price is higher than selling price (negative margin)")
                
                # Add vendor to validated data
                validated_data['vendor'] = business_partner
                
                # If there are errors, skip this row
                if row_errors:
                    results['errors'].extend(row_errors)
                    results['error_count'] += 1
                    continue
                
                # Add warnings to results
                if row_warnings:
                    results['warnings'].extend(row_warnings)
                
                # Create or update part if not validation-only mode
                if not validate_only:
                    compatible_vehicles = validated_data.pop('compatible_vehicles', [])
                    
                    if existing_part:
                        # Update existing part
                        for key, value in validated_data.items():
                            if key != 'vendor':  # Don't update vendor
                                setattr(existing_part, key, value)
                        existing_part.save()
                        
                        # Update vehicle compatibility
                        if compatible_vehicles:
                            existing_part.compatible_vehicles.set(compatible_vehicles)
                        
                        results['updated_count'] += 1
                    else:
                        # Create new part
                        new_part = Part.objects.create(**validated_data)
                        
                        # Set vehicle compatibility
                        if compatible_vehicles:
                            new_part.compatible_vehicles.set(compatible_vehicles)
                        
                        results['created_count'] += 1
                
                results['valid_count'] += 1
                
            except Exception as e:
                error_msg = f"Row {row_num}: Unexpected error - {str(e)}"
                results['errors'].append(error_msg)
                results['error_count'] += 1
        
        # Generate validation summary
        results['validation_summary'] = {
            'total_processed': results['total_rows'],
            'success_rate': (results['valid_count'] / results['total_rows'] * 100) if results['total_rows'] > 0 else 0,
            'most_common_errors': sorted(results['field_errors'].items(), key=lambda x: x[1], reverse=True)[:5]
        }
    
    except Exception as e:
        results['errors'].append(f"File processing error: {str(e)}")
        results['error_count'] += 1
    
    return results


def generate_excel_export(queryset, include_images):
    """
    Generate Excel export of parts data.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        import io
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Vendor Parts"
        
        # Define headers
        headers = [
            'Part Number', 'Name', 'Description', 'Material Type', 'Plant',
            'Material Group', 'Purchasing Group', 'Base Unit of Measure',
            'Purchase Order Unit', 'Order Unit', 'Numerator for Conversion',
            'Denominator for Conversion', 'Gross Weight', 'Net Weight',
            'Weight Unit', 'Volume', 'Volume Unit', 'Size/Dimensions',
            'EAN/UPC', 'Manufacturer Part Number', 'Old Material Number',
            'Planned Delivery Time (Days)', 'Goods Receipt Processing Time (Days)',
            'Valuation Class', 'Price Unit (PEINH)', 'Moving Average Price',
            'Category', 'Brand', 'Price', 'Quantity', 'Safety Stock',
            'Reorder Point', 'Warranty Period', 'Active', 'Featured',
            'Created At', 'Updated At'
        ]
        
        if include_images:
            headers.append('Image URL')
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data rows
        for row_num, part in enumerate(queryset, 2):
            data = [
                part.part_number,
                part.name,
                part.description or '',
                part.material_type or '',
                part.plant or '',
                part.material_group or '',
                part.purchasing_group or '',
                part.base_unit_of_measure or '',
                part.purchase_order_unit or '',
                part.order_unit or '',
                part.numerator_for_conversion or '',
                part.denominator_for_conversion or '',
                part.gross_weight or '',
                part.net_weight or '',
                part.weight_unit or '',
                part.volume or '',
                part.volume_unit or '',
                part.size_dimensions or '',
                part.ean_upc or '',
                part.manufacturer_part_number or '',
                part.old_material_number or '',
                part.planned_delivery_time_days or '',
                part.goods_receipt_processing_time_days or '',
                part.valuation_class or '',
                part.price_unit_peinh or '',
                part.moving_average_price or '',
                part.category.name if part.category else '',
                part.brand.name if part.brand else '',
                float(part.price) if part.price else '',
                int(part.quantity) if part.quantity else '',
                int(part.safety_stock) if part.safety_stock else '',
                int(part.reorder_point) if part.reorder_point else '',
                part.warranty_period or '',
                'Yes' if part.is_active else 'No',
                'Yes' if part.is_featured else 'No',
                part.created_at.strftime('%Y-%m-%d %H:%M:%S') if part.created_at else '',
                part.updated_at.strftime('%Y-%m-%d %H:%M:%S') if part.updated_at else '',
            ]
            
            if include_images:
                image_url = part.image.url if part.image else ''
                data.append(image_url)
            
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="vendor_parts_export.xlsx"'
        
        # Save workbook to response
        workbook.save(response)
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return generate_csv_export(queryset, include_images)


def generate_csv_export(queryset, include_images):
    """
    Generate CSV export of parts data (enhanced version of existing function).
    """
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="vendor_parts_export.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    headers = [
        'Part Number', 'Name', 'Description', 'Material Type', 'Plant',
        'Material Group', 'Purchasing Group', 'Base Unit of Measure',
        'Purchase Order Unit', 'Order Unit', 'Numerator for Conversion',
        'Denominator for Conversion', 'Gross Weight', 'Net Weight',
        'Weight Unit', 'Volume', 'Volume Unit', 'Size/Dimensions',
        'EAN/UPC', 'Manufacturer Part Number', 'Old Material Number',
        'Planned Delivery Time (Days)', 'Goods Receipt Processing Time (Days)',
        'Valuation Class', 'Price Unit (PEINH)', 'Moving Average Price',
        'Category', 'Brand', 'Price', 'Quantity', 'Safety Stock',
        'Reorder Point', 'Warranty Period', 'Active', 'Featured',
        'Created At', 'Updated At'
    ]
    
    if include_images:
        headers.append('Image URL')
    
    writer.writerow(headers)
    
    # Write data rows
    for part in queryset:
        row = [
            part.part_number,
            part.name,
            part.description or '',
            part.material_type or '',
            part.plant or '',
            part.material_group or '',
            part.purchasing_group or '',
            part.base_unit_of_measure or '',
            part.purchase_order_unit or '',
            part.order_unit or '',
            part.numerator_for_conversion or '',
            part.denominator_for_conversion or '',
            part.gross_weight or '',
            part.net_weight or '',
            part.weight_unit or '',
            part.volume or '',
            part.volume_unit or '',
            part.size_dimensions or '',
            part.ean_upc or '',
            part.manufacturer_part_number or '',
            part.old_material_number or '',
            part.planned_delivery_time_days or '',
            part.goods_receipt_processing_time_days or '',
            part.valuation_class or '',
            part.price_unit_peinh or '',
            part.moving_average_price or '',
            part.category.name if part.category else '',
            part.brand.name if part.brand else '',
            part.price or '',
            part.quantity or '',
            part.safety_stock or '',
            part.reorder_point or '',
            part.warranty_period or '',
            'Yes' if part.is_active else 'No',
            'Yes' if part.is_featured else 'No',
            part.created_at.strftime('%Y-%m-%d %H:%M:%S') if part.created_at else '',
            part.updated_at.strftime('%Y-%m-%d %H:%M:%S') if part.updated_at else '',
        ]
        
        if include_images:
            image_url = part.image.url if part.image else ''
            row.append(image_url)
        
        writer.writerow(row)
    
    return response